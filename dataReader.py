from openpyxl import load_workbook
from settings import dataNode, dataFile
from datetime import datetime

def readByIndex(i) -> dataNode:
    wb = load_workbook(dataFile)
    sheet = wb.active

    tempDataNode = dataNode()
    #for i in range(3, 16):
    #    print(i, sheet.cell(row=3, column=i).value)
    tempDataNode.id = i
    tempDataNode.userAgent = str(sheet.cell(row=i, column=3).value)
    tempDataNode.firstName = str(sheet.cell(row=i, column=4).value)
    tempDataNode.secondName = str(sheet.cell(row=i, column=5).value)
    tempDate = [0,0,0]
    try:
        tempDate = str(str(sheet.cell(row=i, column=6).value).strftime("%d.%m.%Y")).split('.')
    except:
        tempDate= str(sheet.cell(row=i, column=6).value).split('.')  # .strftime("%d/%m/%Y")
    tempDataNode.date = tempDate
    tempStr = ''

    if str(sheet.cell(row=i, column=7).value):
        tempStr += str(sheet.cell(row=i, column=7).value)

    if str(sheet.cell(row=i, column=8).value):
        tempStr += ' ' + str(sheet.cell(row=i, column=8).value)

    if str(sheet.cell(row=i, column=9).value):
        tempStr += ' ' + str(sheet.cell(row=i, column=9).value)

    if str(sheet.cell(row=i, column=10).value):
        tempStr += ' ' + str(sheet.cell(row=i, column=10).value)

    tempDataNode.address = tempStr

    tempDataNode.phone = str(sheet.cell(row=i, column=11).value)
    tempDataNode.mail = str(sheet.cell(row=i, column=12).value)

    tempDataNode.login = str(sheet.cell(row=i, column=13).value)
    tempDataNode.password = str(sheet.cell(row=i, column=14).value)
    tempDataNode.answer = str(sheet.cell(row=i, column=15).value)

    return tempDataNode